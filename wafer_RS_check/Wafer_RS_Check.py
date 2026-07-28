import os
import pandas as pd
import zipfile
from tkinter import Tk, Button, filedialog, Text, Scrollbar, END
from datetime import datetime
from openpyxl.styles import Border, Side, PatternFill
import threading
import gc

# ====================== Helpers ======================

def select_zip_files():
    return filedialog.askopenfilenames(title="Select ZIP Files", filetypes=[("ZIP Files", "*.zip")])

def select_output_folder():
    return filedialog.askdirectory(title="Select Output Folder")

def apply_borders(ws):
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

def adjust_column_width(ws):
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_len + 2, 40)
        ws.column_dimensions[letter].width = adjusted_width

# ====================== Core: Process NG Rows with Comment Column ======================

def process_summary_ng_only(input_folder, output_folder, log_terminal):
    try:
        csv_files = [f for f in os.listdir(input_folder) if f.lower().endswith('.csv')]
        if not csv_files:
            log_terminal.insert(END, "⚠️ No CSV files found.\n")
            log_terminal.see(END)
            return

        LSL, USL = 264000, 394000
        all_ng_rows = []

        log_terminal.insert(END, f"🔍 Scanning {len(csv_files)} CSV files for NG rows...\n")
        log_terminal.see(END)

        for idx, csv_file in enumerate(csv_files, 1):
            try:
                df = pd.read_csv(os.path.join(input_folder, csv_file), low_memory=False)

                if "RS_PAD4_2" not in df.columns:
                    continue

                rs_vals = pd.to_numeric(df["RS_PAD4_2"], errors='coerce')
                ng_mask = (rs_vals < LSL) | (rs_vals > USL)
                ng_rows = df[ng_mask].copy()

                if not ng_rows.empty:
                    # Add Comment column
                    ng_rows['Comment'] = ng_rows['RS_PAD4_2'].apply(
                        lambda x: "超上限" if pd.notna(x) and x > USL else ("超下限" if pd.notna(x) and x < LSL else "")
                    )
                    all_ng_rows.append(ng_rows)

                if idx % 10 == 0 or idx == len(csv_files):
                    log_terminal.insert(END, f"  → Checked {idx}/{len(csv_files)} files\n")
                    log_terminal.see(END)

                del df, rs_vals, ng_rows
                gc.collect()

            except Exception:
                continue  # Skip problematic files

        if not all_ng_rows:
            log_terminal.insert(END, "✅ No NG rows found in any CSV file.\n")
            log_terminal.see(END)
            return

        summary_df = pd.concat(all_ng_rows, ignore_index=True)
        log_terminal.insert(END, f"📊 Total NG rows found: {len(summary_df)}\n")
        log_terminal.see(END)

        # Ensure 'Comment' is the last column
        cols = [col for col in summary_df.columns if col != 'Comment'] + ['Comment']
        summary_df = summary_df[cols]

        output_file = os.path.join(output_folder, f"TSMC_DOE_電阻值檢查_Summary-{datetime.now().strftime('%Y%m%d')}.xlsx")
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            summary_df.to_excel(writer, index=False, sheet_name='NG_Summary')
            ws = writer.sheets['NG_Summary']
            apply_borders(ws)
            adjust_column_width(ws)

            # Find column letters
            rs_col_letter = None
            for cell in ws[1]:
                if cell.value == "RS_PAD4_2":
                    rs_col_letter = cell.column_letter
                    break

            # Highlight RS_PAD4_2 cells in red
            if rs_col_letter:
                for row in range(2, ws.max_row + 1):
                    ws[f"{rs_col_letter}{row}"].fill = red_fill

        log_terminal.insert(END, f"✅ Summary saved: {output_file}\n")
        log_terminal.see(END)

    except Exception as e:
        log_terminal.insert(END, f"💥 Error: {str(e)}\n")
        log_terminal.see(END)

# ====================== Extract CSV from ZIP ======================

def extract_csv_from_zip(zip_files, output_folder, log_terminal):
    extracted = 0
    for zip_path in zip_files:
        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                for name in zf.namelist():
                    if name.lower().endswith('.csv'):
                        filename = os.path.basename(name)
                        if filename:
                            dest = os.path.join(output_folder, filename)
                            base, ext = os.path.splitext(filename)
                            counter = 1
                            while os.path.exists(dest):
                                dest = os.path.join(output_folder, f"{base}_{counter}{ext}")
                                counter += 1
                            with zf.open(name) as src, open(dest, 'wb') as dst:
                                dst.write(src.read())
                            extracted += 1
        except Exception as e:
            log_terminal.insert(END, f"⚠️ Failed to extract {zip_path}: {str(e)}\n")
            log_terminal.see(END)
    log_terminal.insert(END, f"✅ Extracted {extracted} CSV file(s).\n")
    log_terminal.see(END)

# ====================== GUI ======================

def main():
    root = Tk()
    root.title("RS_PAD4_2 NG Summary - RED HIGHLIGHT + Comment")
    w, h = 620, 500
    x = (root.winfo_screenwidth() - w) // 2
    y = (root.winfo_screenheight() - h) // 2
    root.geometry(f"{w}x{h}+{x}+{y}")

    zip_files = []
    output_folder = ""

    log = Text(root, wrap="word", bg="black", fg="#ff9999", font=("Consolas", 10))
    log.pack(pady=10, fill="both", expand=True)
    scroll = Scrollbar(log)
    scroll.pack(side="right", fill="y")
    log.config(yscrollcommand=scroll.set)
    scroll.config(command=log.yview)

    def pick_zips():
        nonlocal zip_files
        zip_files = select_zip_files()
        log.insert(END, f"📁 Selected {len(zip_files)} ZIP(s)\n")
        log.see(END)

    def pick_out():
        nonlocal output_folder
        output_folder = select_output_folder()
        log.insert(END, f"📂 Output: {output_folder}\n" if output_folder else "⚠️ No output folder\n")
        log.see(END)

    def run():
        if not zip_files or not output_folder:
            log.insert(END, "⚠️ Please select ZIP files and output folder\n")
            log.see(END)
            return

        def task():
            log.insert(END, "🔄 Starting extraction and analysis...\n")
            log.see(END)
            extract_csv_from_zip(zip_files, output_folder, log)
            process_summary_ng_only(output_folder, output_folder, log)
            gc.collect()

        threading.Thread(target=task, daemon=True).start()

    def clear():
        log.delete(1.0, END)

    Button(root, text="Select ZIP Files", command=pick_zips, width=25).pack(pady=2)
    Button(root, text="Select Output Folder", command=pick_out, width=25).pack(pady=2)
    Button(root, text="Run Analysis", command=run, bg="#4CAF50", fg="white", width=25).pack(pady=5)
    Button(root, text="Clear Log", command=clear, width=25).pack(pady=2)

    root.mainloop()

# ====================== Entry Point ======================

if __name__ == "__main__":
    main()
